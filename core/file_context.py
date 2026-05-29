"""
Expand local file references inside Swarm prompts so employee CLIs receive real content.

Markers (in CONTEXT or TASK, before ``engine.run``):

- ``[[swarm-file:relative/or/absolute/path]]`` — inline a text preview, image (Markdown), PDF text, or binary digest.
- Lines starting with ``INCLUDE_FILE: path`` — same as a marker at that line.

Optional: ``auto_backtick_paths`` in config — existing files whose path appears in backticks
`` `like/this.png` `` are expanded once (use with care).

Environment overrides: ``SWARM_FILE_CONTEXT`` (0 to disable), ``SWARM_FILE_CONTEXT_MAX_TOTAL``,
``SWARM_FILE_CONTEXT_MAX_TEXT``, ``SWARM_FILE_CONTEXT_MAX_IMAGE_B64``.

Optional: ``pip install pypdf`` for better PDF text extraction.
"""

from __future__ import annotations

import base64
import hashlib
import json
import os
import re
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

SWARM_HOME = Path(os.environ.get("SWARM_HOME", Path.home() / ".swarm"))
IMAGE_CACHE = SWARM_HOME / "prompt-images"

MARKER_RE = re.compile(r"\[\[swarm-file:\s*([^\]\n]+?)\s*\]\]", re.IGNORECASE)
INCLUDE_RE = re.compile(r"^INCLUDE_FILE:\s*(.+?)\s*$", re.IGNORECASE | re.MULTILINE)
BACKTICK_RE = re.compile(r"`([^`\n][^`]{0,800})`")


_TEXT_EXT = {
    ".txt",
    ".md",
    ".mdx",
    ".markdown",
    ".json",
    ".jsonc",
    ".yaml",
    ".yml",
    ".toml",
    ".csv",
    ".tsv",
    ".xml",
    ".html",
    ".htm",
    ".css",
    ".scss",
    ".sass",
    ".less",
    ".js",
    ".jsx",
    ".ts",
    ".tsx",
    ".mjs",
    ".cjs",
    ".vue",
    ".svelte",
    ".py",
    ".pyw",
    ".pyi",
    ".rb",
    ".php",
    ".java",
    ".kt",
    ".kts",
    ".go",
    ".rs",
    ".c",
    ".h",
    ".cc",
    ".cpp",
    ".hpp",
    ".swift",
    ".m",
    ".mm",
    ".sh",
    ".bash",
    ".zsh",
    ".fish",
    ".ps1",
    ".bat",
    ".cmd",
    ".cmake",
    ".make",
    ".mk",
    "dockerfile",
    ".dockerignore",
    ".gitignore",
    ".env.example",
    ".graphql",
    ".sql",
    ".r",
    ".lua",
    ".pl",
    ".pm",
    ".ex",
    ".exs",
    ".erl",
    ".cljs",
    ".cljc",
    ".edn",
    ".tex",
    ".rst",
    ".ini",
    ".cfg",
    ".conf",
    ".properties",
    ".svg",
}


_IMAGE_EXT = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".ico"}


def _env_int(key: str, default: int) -> int:
    raw = os.environ.get(key)
    if raw is None or raw.strip() == "":
        return default
    try:
        return int(raw)
    except ValueError:
        return default


def _effective_options(user: dict[str, Any] | None) -> dict[str, Any]:
    defaults = {
        "enabled": True,
        "max_total_chars": _env_int("SWARM_FILE_CONTEXT_MAX_TOTAL", 220_000),
        "max_text_chars_per_file": _env_int("SWARM_FILE_CONTEXT_MAX_TEXT", 100_000),
        "max_image_base64_chars": _env_int("SWARM_FILE_CONTEXT_MAX_IMAGE_B64", 380_000),
        "auto_backtick_paths": False,
    }
    if not user:
        return defaults
    out = dict(defaults)
    for k, v in user.items():
        if k == "enabled" and v is False:
            out["enabled"] = False
        elif k == "enabled" and v is True:
            out["enabled"] = True
        elif k in defaults and v is not None:
            out[k] = v
    if os.environ.get("SWARM_FILE_CONTEXT", "").strip().lower() in {"0", "false", "no", "off"}:
        out["enabled"] = False
    return out


def _normalize_roots(paths: list[Path | str] | None) -> list[Path]:
    roots: list[Path] = []
    for p in paths or []:
        try:
            roots.append(Path(p).expanduser().resolve())
        except (OSError, ValueError):
            continue
    # de-dupe keep order
    seen: set[Path] = set()
    unique: list[Path] = []
    for r in roots:
        if r not in seen:
            unique.append(r)
            seen.add(r)
    return unique


def _resolve_path(spec: str, roots: list[Path]) -> Path | None:
    raw = spec.strip().strip("\"'").strip()
    if not raw or raw.startswith("-"):
        return None
    p = Path(raw).expanduser()
    if p.is_absolute():
        try:
            rp = p.resolve()
            return rp if rp.is_file() else None
        except OSError:
            return None
    for root in roots:
        try:
            cand = (root / raw).resolve()
            if cand.is_file():
                return cand
        except OSError:
            continue
    try:
        cwd_hit = (Path.cwd() / raw).resolve()
        if cwd_hit.is_file():
            return cwd_hit
    except OSError:
        pass
    return None


def _mime_for_image(path: Path) -> str:
    ext = path.suffix.lower()
    mapping = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".gif": "image/gif",
        ".webp": "image/webp",
        ".bmp": "image/bmp",
        ".ico": "image/x-icon",
    }
    return mapping.get(ext, "application/octet-stream")


def _image_markdown(path: Path, max_b64: int) -> str:
    try:
        data = path.read_bytes()
    except OSError as e:
        return f"_(cannot read image: {e})_"
    if not data:
        return "_(empty image)_"
    b64 = base64.standard_b64encode(data).decode("ascii")
    mime = _mime_for_image(path)
    n = len(b64)
    if max_b64 > 0 and n <= max_b64:
        return (
            f"### Image: `{path}`\n\n"
            f"![included file `{path.name}`](data:{mime};base64,{b64})\n\n"
            f"*(MIME `{mime}`, {len(data)} bytes, {n} base64 chars — model should describe and use visual detail.)*"
        )
    IMAGE_CACHE.mkdir(parents=True, exist_ok=True)
    digest = hashlib.sha256(data).hexdigest()[:16]
    ext = path.suffix or ".bin"
    out = IMAGE_CACHE / f"{_safe_fname(path.name)}__{digest}{ext}"
    out.write_bytes(data)
    uri = out.resolve().as_uri()
    return (
        f"### Image (spooled): `{path}`\n\n"
        f"![included file `{path.name}`]({uri})\n\n"
        f"`{out}` _(large — {len(data)} bytes; open path or tune `SWARM_FILE_CONTEXT_MAX_IMAGE_B64`)_"
    )


def _safe_fname(s: str) -> str:
    return re.sub(r"[^a-zA-Z0-9._-]+", "_", s)[:100] if s else "file"


def _text_snippet(label: Path, raw: str, limit: int) -> str:
    if len(raw) > limit:
        raw = raw[:limit] + "\n\n_(…truncated)_"
    return f"### Text file: `{label}`\n\n```\n{raw}\n```\n"


def _try_read_plain_text(path: Path, limit: int) -> tuple[str | None, str | None]:
    try:
        data = path.read_bytes()
    except OSError as e:
        return None, str(e)
    if b"\x00" in data[:8192]:
        return None, "binary"
    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError:
        try:
            text = data.decode("latin-1")
        except UnicodeDecodeError:
            return None, "undecodable"
    if len(text) > limit:
        text = text[:limit] + "\n\n_(…truncated)_"
    return text, None


def _pdf_extract(path: Path, limit: int) -> str | None:
    try:
        from pypdf import PdfReader  # type: ignore[import-untyped]
    except ImportError:
        return None
    try:
        reader = PdfReader(str(path))
        parts: list[str] = []
        for page in reader.pages[:80]:
            t = page.extract_text()
            if t:
                parts.append(t)
        blob = "\n\n".join(parts).strip()
        if not blob:
            return None
        return blob[:limit] + ("\n\n_(…truncated)_" if len(blob) > limit else "")
    except Exception:
        return None


def _docx_quick_text(path: Path, limit: int) -> str | None:
    try:
        with zipfile.ZipFile(path, "r") as zf:
            with zf.open("word/document.xml") as xf:
                tree = ET.parse(xf)
    except Exception:
        return None
    try:
        ns = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
        texts = [node.text or "" for node in tree.iter(f"{ns}t")]
        blob = "".join(texts).strip()
        if not blob:
            return None
        return blob[:limit] + ("\n\n_(…truncated)_" if len(blob) > limit else "")
    except Exception:
        return None


def _xlsx_first_sheet_cells(path: Path, limit: int) -> str | None:
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError:
        return None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        sheet = wb.active
        rows_: list[list[str]] = []
        for row in sheet.iter_rows(max_row=500, values_only=True):
            cells = [(str(c) if c is not None else "") for c in row]
            if any(cell.strip() for cell in cells):
                rows_.append(cells)
        wb.close()
        if not rows_:
            return None
        lines = ["\t".join(r) for r in rows_[:400]]
        blob = "\n".join(lines)
        return blob[:limit] + ("\n\n_(…truncated)_" if len(blob) > limit else "")
    except Exception:
        return None


def _binary_digest(path: Path, read_cap: int = 48_576) -> str:
    try:
        size = path.stat().st_size
    except OSError as e:
        return f"_(stat failed: {e})_"
    try:
        with path.open("rb") as f:
            head = f.read(read_cap)
    except OSError as e:
        return f"_(read failed: {e})_"
    hexlen = len(head) * 2
    preview = head[:256].hex()
    return (
        f"### Binary file: `{path}`\n\n"
        f"- Size: **{size}** bytes\n"
        f"- First **{len(head)}** bytes as hex ({hexlen} chars): `{preview}{'…' if len(head) > 256 else ''}`\n"
        f"- _Desk should request conversion or MCP tools if it needs semantic content._\n"
    )


def inline_file(path: Path, opts: dict[str, Any]) -> str:
    ext = path.suffix.lower()
    name_lower = path.name.lower()
    max_text = int(opts["max_text_chars_per_file"])
    max_img = int(opts["max_image_base64_chars"])

    if ext in _IMAGE_EXT or (ext == "" and magic_guess(path) == "image"):
        return _image_markdown(path, max_img)

    if ext == ".pdf" or name_lower.endswith(".pdf"):
        t = _pdf_extract(path, max_text)
        if t:
            return f"### PDF (extracted text): `{path}`\n\n```\n{t}\n```\n"
        return _binary_digest(path) + "\n_(Install `pypdf` for text extraction.)_\n"

    if ext == ".docx":
        t = _docx_quick_text(path, max_text)
        if t:
            return f"### DOCX (plain text approximation): `{path}`\n\n```\n{t}\n```\n"
        return _binary_digest(path) + "\n_(DOCX text extraction failed or empty — inspect with proper tools.)_\n"

    if ext == ".xlsx":
        t = _xlsx_first_sheet_cells(path, max_text)
        if t:
            return f"### XLSX (first sheet preview): `{path}`\n\n```\n{t}\n```\n"
        return (
            "### XLSX: `{path}`\n\n"
            "`openpyxl` not installed — install for tabular preview, or inspect via scripts.\n"
        ).format(path=path)

    if ext in _TEXT_EXT or name_lower == "dockerfile" or name_lower.startswith("readme."):
        if ext == ".json":
            try:
                parsed = json.loads(path.read_text(encoding="utf-8"))
                blob = json.dumps(parsed, indent=2, ensure_ascii=False)
                if len(blob) > max_text:
                    blob = blob[:max_text] + "\n\n_(…truncated)_"
                return f"### JSON file: `{path}`\n\n```json\n{blob}\n```\n"
            except (OSError, UnicodeDecodeError, json.JSONDecodeError):
                text, err = _try_read_plain_text(path, max_text)
                if text is None:
                    return _binary_digest(path) if err == "binary" else f"_(unreadable json: `{path}`)_"
                return _text_snippet(path, text, max_text)

        text, err = _try_read_plain_text(path, max_text)
        if text is not None:
            return _text_snippet(path, text, max_text)
        if err == "binary":
            return _binary_digest(path)

    # default: attempt utf-8; else binary digest
    text, err = _try_read_plain_text(path, max_text)
    if text is not None:
        return _text_snippet(path, text, max_text)
    return _binary_digest(path)


def magic_guess(path: Path) -> str:
    """Very small sniff for extensionless paths."""
    try:
        buf = path.read_bytes()[:16]
    except OSError:
        return "unknown"
    if buf.startswith(b"\x89PNG\r\n\x1a\n"):
        return "image"
    if buf[:3] == b"\xff\xd8\xff":
        return "image"
    if buf.startswith(b"GIF87a") or buf.startswith(b"GIF89a"):
        return "image"
    if buf.startswith(b"RIFF") and len(buf) >= 12 and buf[8:12] == b"WEBP":
        return "image"
    return "unknown"


_BUDGET: dict[str, int] = {}


def _budget_take(n: int, opts: dict[str, Any]) -> bool:
    cap = int(opts["max_total_chars"])
    if cap <= 0:
        return False
    used = _BUDGET.get("used", 0)
    if used + n > cap:
        return False
    _BUDGET["used"] = used + n
    return True


def _expand_one(spec: str, roots: list[Path], opts: dict[str, Any]) -> str:
    path = _resolve_path(spec, roots)
    if path is None:
        tried = ", ".join(str(r) for r in roots[:6])
        more = f" (+{len(roots) - 6} more)" if len(roots) > 6 else ""
        return f"_(**swarm-file** not found: `{spec.strip()}` — searched under: {tried}{more})_"
    if not path.is_file():
        return f"_(**swarm-file** not a regular file: `{path}`)_"
    snippet = inline_file(path, opts)
    n = len(snippet)
    if not _budget_take(n, opts):
        return f"_(**swarm-file** skipped `{path}`: total CONTEXT budget {_BUDGET.get('used', 0)} / {opts['max_total_chars']} chars)_"
    return snippet


def expand_prompt_file_references(prompt: str, *, search_roots: list[Path | str] | None, config_section: dict | bool | None) -> str:
    """
    Replace ``[[swarm-file:…]]`` and ``INCLUDE_FILE:`` markers with inlined previews.
    """
    global _BUDGET
    if config_section is False:
        return prompt
    user = config_section if isinstance(config_section, dict) else {}
    opts = _effective_options(user)
    if not opts["enabled"]:
        return prompt

    _BUDGET = {"used": 0}
    roots = _normalize_roots(search_roots)
    if not roots:
        roots = [Path.cwd().resolve()]

    out = prompt

    def repl_marker(m: re.Match[str]) -> str:
        return _expand_one(m.group(1), roots, opts)

    out = MARKER_RE.sub(repl_marker, out)

    def repl_include(m: re.Match[str]) -> str:
        body = _expand_one(m.group(1), roots, opts)
        return body

    out = INCLUDE_RE.sub(repl_include, out)

    if opts.get("auto_backtick_paths"):
        seen_paths: set[str] = set()

        def maybe_backtick(m: re.Match[str]) -> str:
            inner = m.group(1).strip()
            if "/" not in inner and "\\" not in inner and "." not in inner:
                return m.group(0)
            if inner in seen_paths:
                return m.group(0)
            cand = _resolve_path(inner, roots)
            if cand is None or not cand.is_file():
                return m.group(0)
            seen_paths.add(inner)
            return _expand_one(inner, roots, opts)

        out = BACKTICK_RE.sub(maybe_backtick, out)

    return out


def file_context_instructions() -> str:
    """Short hint for CONTEXT (optional prepend)."""
    return (
        "## Reading local files inside Swarm\n\n"
        "Boss may attach file contents inline using:\n\n"
        "- `[[swarm-file:relative/or/absolute/path]]` anywhere in CONTEXT or TASK\n"
        "- a line `INCLUDE_FILE: path/to/file`\n\n"
        "Images are embedded for vision-capable desk CLIs (Markdown `data:` or spooled paths). "
        "PDF/DOCX/XLSX/text/code are pasted as excerpts; binaries get a structured hex digest.\n"
    )
